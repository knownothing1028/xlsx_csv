import csv
import openpyxl
from openpyxl import Workbook
from xlrd import sheet
from openpyxl import load_workbook

# get the csv no indent title list
csvfile = []
with open('report.csv') as f:
	reader = csv.reader(f)
	for row in reader:
		csvfile.append(row)

csv_title = []

for t in csvfile[0]:
	csv_title.append(t.replace(' ','').replace('_','').lower())

# get xlsx no indent title list
wb = load_workbook('whatever.xlsx')
sheet = wb.active
max_col = sheet.max_column

xlsx_tag = []

for i in range(1,max_col):
	xlsx_tag.append(sheet.cell(1,i).value.replace(' ','').replace('_','').lower())

# get the compatitable
compat_dct ={}
compat_list = []
for i in csv_title:
	if i in xlsx_tag:
		compat_dct[i] = [csv_title.index(i),xlsx_tag.index(i)]
		compat_list.append([csv_title.index(i),xlsx_tag.index(i)])

print(csv_title)
print(xlsx_tag)


print(compat_dct)
# get data from csv by column
# write it into xlsx by column
row = 1
for i in csvfile:
	if row == 1:
		row += 1
		continue

	else:
		for pos in compat_list:
			sheet.cell(row = row, column= pos[1]+1).value = i[pos[0]]
		row +=1

wb.save('whatever.xlsx')




