from tkinter import messagebox

from openpyxl import load_workbook
from tkinter import *

tag_format = ['no.', 'testround', 'devicename', 'casename', 'result', 'manual', 'comments', 'errortype', 'errorscene', 'time(minute)', 'testplanname', 'location', 'devicetype', 'csmbuildversion', 'nodedbtversion', 'executiondate', 'tester', 'failedreason', 'bugid', 'automationprecondition', 'automationcomments','tag1']
tag_need_fill = ['testround','devicename','time(minute)','location','devicetype','nodedbtversion','automationprecondition','tag1']

path_format = 'format.xlsx'
path1 = 'MY21GA_SIGNED_0925_yichi_zhang.xlsx'
path2 = 'MY21GA_SIGNED_0926_yichi_zhang.xlsx'
path3 = 'MY21GA_SIGNED_0930_yichi_zhang.xlsx'

class XLSXfile():
	path = ''
	list_tag = []
	max_col = 0
	max_row = 0
	wb = None
	sheet = None

	def __init__(self,path):
		self.path = path
		self.wb  = load_workbook(path)
		self.sheet  = self.wb.active
		self.max_col = self.sheet.max_column
		self.max_row = self.sheet.max_row

		if(self.max_col>0):
			for i in range(1,self.max_col+1):
				self.list_tag.append(self.sheet.cell(1,i).value)

	def find_last_row(self):
		row = 0
		endrow = 0
		while(row< self.max_row):
			row +=1
			if(self.sheet.cell(row,4).value is None):
				continue
			else:
				endrow = row
		self.max_row = endrow


	def format_uniform(self):
		list_uniform = []
		for i in self.list_tag:
			list_uniform.append(i.replace(' ', '').replace('_', '').lower())
		self.list_tag = list_uniform
		return list_uniform

	def tag_compare(self, data_list_tag):
		list_relation = []
		self.format_uniform()
		for i in self.list_tag:
			if i in data_list_tag:
				list_relation.append([i, self.list_tag.index(i), data_list_tag.index(i)])
			else:
				print("No such data", i)
				continue
		return list_relation

	def data_transfer(self,data_sheet,list_relation,data_rows):
		self.find_last_row()
		print('source rows',self.max_row)
		count  = 0
		print("start transfer","we have max_row",data_rows)
		for j in range(2,data_rows+1):
			print("lines = ",j)
			for i in list_relation:
				# print(self.sheet.cell(self.max_row+j-1,i[1]+1).value,data_sheet.cell(j,i[2]+1).value)
				self.sheet.cell(self.max_row+j-1,i[1]+1).value = data_sheet.cell(j,i[2]+1).value
			count +=1

		self.max_row+=count
		print("afterall the last_row is", self.max_row)
		self.wb.save(self.path)
		print('Combine',count,'records!')


	def edit_no(self):
		self.format_uniform()
		self.find_last_row()
		no_index = self.list_tag.index('no.')
		for i in range(1,self.max_row):
			print(i,end=",")
			print(self.sheet.cell(i+1,no_index+1).value)
			self.sheet.cell(i+1,no_index+1).value = i
		self.wb.save(self.path)

	def need_auto_fill(self):
		root = Tk()
		root.title("Cell need fill")
		root.geometry('300x500')

		value_tag_need_fill = []
		value = []
		for i in tag_need_fill:
			item = Label(root,text = i)
			item.pack()
			item_text = StringVar()
			item_object = Entry(root, textvariable =item_text)
			value_tag_need_fill.append(item_text)
			item_text.set("")
			item_object.pack()
		count = 0


		def on_click():
			for i in value_tag_need_fill:
				value.append(i.get())
			string = ''
			for i in range(0, len(tag_need_fill)):
				string = string + tag_need_fill[i] + ':' + value[i] + "\n"
			print(string)
			messagebox.showinfo(title='Value Written', message=string)
			root.quit()

		Button(root,text = 'Ready to Write',command = on_click).pack()
		root.mainloop()

		for tag in tag_need_fill:
			for i in range(1, self.max_row):
				self.sheet.cell(i+1, self.list_tag.index(tag)+1).value = value[tag_need_fill.index(tag)]

		self.wb.save(self.path)
		print("fill done!!!")


# just for testing

xlsx_format = XLSXfile(path_format)
xlsx_format.format_uniform()
xlsx_format.find_last_row()
xlsx_format.edit_no()
xlsx_format.need_auto_fill()
# xlsx_data1 = XLSXfile(path1)
# xlsx_data1.format_uniform()
# list_relation = xlsx_format.tag_compare(xlsx_data1.list_tag)
# xlsx_data1.find_last_row()
# print("1max_row", xlsx_data1.max_row)
# xlsx_format.data_transfer(xlsx_data1.sheet,list_relation,xlsx_data1.max_row)
#
# xlsx_data2 = XLSXfile(path2)
# xlsx_data2.format_uniform()
# list_relation = xlsx_format.tag_compare(xlsx_data2.list_tag)
# xlsx_data2.find_last_row()
# print("2max_row", xlsx_data2.max_row)
# xlsx_format.data_transfer(xlsx_data2.sheet,list_relation,xlsx_data2.max_row)
#
# xlsx_data3 = XLSXfile(path3)
# xlsx_data3.format_uniform()
# list_relation = xlsx_format.tag_compare(xlsx_data3.list_tag)
# xlsx_data3.find_last_row()
# print("3max_row", xlsx_data3.max_row)
# xlsx_format.data_transfer(xlsx_data3.sheet,list_relation,xlsx_data3.max_row)
